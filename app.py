import io
import json
import math
import re
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# DG Declaration Builder
# Python conversion of the supplied HTML application.
# Workflow/field names are intentionally preserved.
# ============================================================

st.set_page_config(page_title="DG Declaration Builder", layout="wide")

st.markdown("""
<style>
:root{
  --ink:#0B1F33; --ink-soft:#33475C; --paper:#F5F2EA;
  --paper-raised:#FFFFFF; --line:#D9D2C2; --placard-red:#B5121B;
}
.stApp { background:#F5F2EA; color:#0B1F33; }
.dg-header {
  background:#0B1F33; color:#F5F2EA; padding:22px 28px 20px;
  border-bottom:4px solid #B5121B; margin-bottom:25px;
}
.dg-header h1 { margin:0; font-size:24px; }
.dg-header p { margin:3px 0 0; color:#B9C4CE; }
</style>
<div class="dg-header">
  <h1>◆ DG Declaration Builder</h1>
  <p>Annexure → verified → IMO Dangerous Goods Declaration (DGDI)</p>
</div>
""", unsafe_allow_html=True)


# ============================= CONSTANTS =============================

CHECKPOINTS = [
    {"key": "flash", "label": "Flash point", "cond": "< 23°C / ≥ 23°C"},
    {"key": "psn", "label": "PSN name", "cond": "Unique proper shipping name"},
    {"key": "pkgcode", "label": "Packing code", "cond": "1st chars: 4G / 1A1 / 1A2"},
    {"key": "mp", "label": "Marine pollutant", "cond": "YES / NO"},
    {"key": "un", "label": "UN number", "cond": "Unique"},
]

FIELD_SYNONYMS = {
    "materialDescription": ["materialdescription","description","productdescription","itemdescription"],
    "unNumber": ["unnumber","unno"],
    "psn": ["unpropershippingname","propershippingname","shippingname","psn"],
    "hazClass": ["hazclass","class","hazardclass","transporthazardclass"],
    "pkgGroup": ["pkggroup","packinggroup","packgroup","unpackinggroup"],
    "marinePollutant": ["marinepollutant","mp"],
    "flashPoint": ["flashpoint","fp"],
    "unCertNo1": ["uncertificateno1","uncertno1"],
    "unCertNoShort": ["uncertificateno","uncertno"],
    "noOfBoxes": ["noofboxesdrums","noofboxes","outerpackages"],
    "noInTins": ["nointins","noinctins","noinstins","notins","innerpackages","noincltins"],
    "grossWt": ["grosswtwithpelletwtinkg","grosswtinkg","grosswt"],
    "netWt": ["netwtinkg","netwt"],
    "ems": ["ems"],
    "technicalName": ["technicalname"],
    "productCode": ["productcode"],
}

REQUIRED_FIELDS = [
    "materialDescription","unNumber","psn","hazClass","pkgGroup",
    "marinePollutant","flashPoint","unCertNo1","noOfBoxes","noInTins",
    "grossWt","netWt"
]

MARKING_RE = re.compile(
    r"([0-9][A-Z][0-9A-Z]?/[A-Z])\s*([0-9*]+/S/[0-9A-Z*]+)\s*(IND/(?:[A-Z]/)?[0-9]+)",
    re.I
)
VALID_RE = re.compile(
    r"valid\s*upto\s*[:\-]?\s*(\d{1,2}\s+[A-Za-z]+\s+\d{4})",
    re.I
)


# ============================= HELPERS =============================

def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def extract_number(s):
    m = re.search(r"-?\d+(?:\.\d+)?", str(s or ""))
    return float(m.group(0)) if m else None


def js_like_round_2(value):
    # Explicit 2-decimal rounding for DG weights.
    # Avoids binary floating-point display such as 757.52 for 757.525.
    return math.floor(value * 100 + 0.5) / 100


def fmt_weight(value):
    return f"{js_like_round_2(value):.2f} KGS"


def uniq(arr):
    return list(dict.fromkeys(
        str(x).strip() for x in arr if str(x).strip()
    ))


def sum_values(arr):
    return sum(float(x or 0) for x in arr)


def match_columns(headers):
    normed = [norm(h) for h in headers]
    used = set()
    mapping = {}

    for field, syns in FIELD_SYNONYMS.items():
        best = -1
        for syn in syns:
            for i, h in enumerate(normed):
                if h == syn and i not in used:
                    best = i
                    break
            if best != -1:
                break

        if best == -1:
            for syn in syns:
                for i, h in enumerate(normed):
                    if syn in h and i not in used:
                        best = i
                        break
                if best != -1:
                    break

        if best != -1:
            mapping[field] = best
            used.add(best)

    return mapping


def pick_best_sheet(xls):
    best = {"name": None, "score": -1, "headerRow": -1, "map": {}}

    for sheet_name in xls.sheet_names:
        aoa = xls.parse(sheet_name, header=None).where(
            pd.notna(xls.parse(sheet_name, header=None)), None
        ).values.tolist()

        for r in range(min(len(aoa), 5)):
            headers = ["" if x is None else str(x) for x in aoa[r]]
            mapping = match_columns(headers)
            score = sum(1 for f in REQUIRED_FIELDS if f in mapping)

            if score > best["score"]:
                best = {
                    "name": sheet_name,
                    "score": score,
                    "headerRow": r,
                    "map": mapping
                }

    return best


def parse_annexure(file_bytes):
    xls = pd.ExcelFile(io.BytesIO(file_bytes), engine=None)
    best = pick_best_sheet(xls)

    raw = xls.parse(best["name"], header=None)
    data_rows = raw.iloc[best["headerRow"] + 1:].values.tolist()
    mapping = best["map"]

    rows = []

    def get(row, field):
        idx = mapping.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    for r in data_rows:
        material = get(r, "materialDescription")
        un_number = get(r, "unNumber")

        if material is None or un_number is None:
            continue
        if str(material).strip().upper() == "TOTAL":
            continue
        if not str(material).strip():
            continue

        def txt(field):
            v = get(r, field)
            return "" if v is None else str(v).strip()

        def num(field):
            v = get(r, field)
            try:
                return float(v)
            except Exception:
                return 0.0

        rows.append({
            "materialDescription": txt("materialDescription"),
            "unNumber": txt("unNumber"),
            "psn": txt("psn"),
            "hazClass": txt("hazClass"),
            "pkgGroup": txt("pkgGroup"),
            "marinePollutant": txt("marinePollutant").upper(),
            "flashPoint": txt("flashPoint"),
            "unCertNo1": txt("unCertNo1"),
            "unCertNoShort": txt("unCertNoShort"),
            "noOfBoxes": num("noOfBoxes"),
            "noInTins": num("noInTins"),
            "grossWt": num("grossWt"),
            "netWt": num("netWt"),
            "ems": txt("ems"),
            "technicalName": txt("technicalName"),
            "productCode": txt("productCode"),
        })

    return {
        "rows": rows,
        "sheetName": best["name"],
        "colMap": mapping,
        "matchedCount": best["score"],
        "totalRequired": len(REQUIRED_FIELDS)
    }


# ============================= PDF TEXT EXTRACTION =============================

def extract_pdf_text(file_bytes):
    text_parts = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text_parts.append(page.extract_text() or "")
    return "\n".join(text_parts) + "\n"


# ============================= MSDS VERIFICATION =============================

def extract_msds_identifiers(text):
    ids = set()

    m = re.search(
        r"Product identifier\s*:?\s*\n?([A-Za-z0-9\-]{2,20})",
        text, re.I
    )
    if m:
        ids.add(m.group(1).strip())

    m = re.search(
        r"Other means of[\s\S]{0,20}identification\s*:?\s*\n?([0-9A-Za-z;,\s]{4,60})",
        text, re.I
    )
    if m:
        for tok in re.split(r"[;,]", m.group(1)):
            tok = tok.strip()
            if tok:
                ids.add(tok)

    return list(ids)


def match_msds_to_row(msds_text, row):
    ids = extract_msds_identifiers(msds_text)
    ids_norm = [norm(x) for x in ids]
    desc_norm = norm(row["materialDescription"])

    for idn in ids_norm:
        if not idn:
            continue
        if idn in desc_norm or norm(row["productCode"]) == idn:
            return True

    return False


def get_section14_block(text):
    m = re.search(r"Section\s*14", text, re.I)
    if not m:
        return text

    after = text[m.start():]
    end = re.search(r"Section\s*1[56]", after, re.I)
    return after[:end.start()] if end else after


def verify_msds_against_row(msds_text, row):
    block = get_section14_block(msds_text)
    block_upper = block.upper()
    checks = []

    digits = re.sub(r"\D", "", row["unNumber"])
    un_re = re.compile(r"UN\s?0*" + re.escape(digits) + r"\b", re.I)
    checks.append({
        "field": "UN Number",
        "ok": bool(un_re.search(block)),
        "expected": row["unNumber"]
    })

    psn_ok = bool(row["psn"]) and row["psn"].upper() in block_upper
    checks.append({
        "field": "Proper shipping name",
        "ok": psn_ok,
        "expected": row["psn"]
    })

    cls_token = re.sub(r"\s", "", row["hazClass"])
    try:
        cls_re = re.compile(
            r"(^|[^0-9.])" + re.escape(cls_token) + r"([^0-9]|$)"
        )
        cls_ok = bool(cls_re.search(block))
    except Exception:
        cls_ok = False

    checks.append({
        "field": "Hazard class",
        "ok": cls_ok,
        "expected": row["hazClass"]
    })

    pg_ok = bool(row["pkgGroup"]) and bool(
        re.search(r"\b" + re.escape(row["pkgGroup"].upper()) + r"\b", block_upper)
    )
    checks.append({
        "field": "Packing group",
        "ok": pg_ok,
        "expected": row["pkgGroup"]
    })

    mp_expected = "YES" if row["marinePollutant"] == "YES" else "NO"
    mp_ok = bool(re.search(r"\bYES\b", block_upper)) if mp_expected == "YES" else bool(
        re.search(r"\bNO\.?\b", block_upper)
    )
    checks.append({
        "field": "Marine pollutant",
        "ok": mp_ok,
        "expected": mp_expected
    })

    fp_matches = re.findall(
        r"Flash point\s*:?\s*([\-]?\d+(?:\.\d+)?)",
        msds_text, re.I
    )
    fp_msds = float(fp_matches[0]) if fp_matches else None
    fp_row = extract_number(row["flashPoint"])
    fp_ok = (
        fp_msds is not None and fp_row is not None
        and abs(fp_msds - fp_row) <= 1.0
    )

    checks.append({
        "field": "Flash point",
        "ok": fp_ok,
        "expected": row["flashPoint"] + (
            f" (MSDS: {fp_msds}°C)" if fp_msds is not None
            else " (not found in MSDS)"
        )
    })

    return checks


# ============================= UN CERTIFICATE VERIFICATION =============================

def extract_cert_info(text):
    m = MARKING_RE.search(text)
    if not m:
        return None

    full = re.sub(r"\s+", " ", f"{m.group(1)} {m.group(2)} {m.group(3)}").strip()
    digits_match = re.search(r"(\d+)$", m.group(3))
    digits = digits_match.group(1) if digits_match else ""
    short_cert = digits[-4:]

    valid_match = VALID_RE.search(text)
    valid_upto = valid_match.group(1) if valid_match else None

    return {
        "full": full,
        "shortCert": short_cert,
        "validUpto": valid_upto
    }


def verify_cert_against_row(cert_info, row):
    checks = []

    marking_ok = (
        cert_info["full"].upper().replace(" ", "")
        == row["unCertNo1"].upper().replace(" ", "")
    )
    checks.append({
        "field": "UN marking string",
        "ok": marking_ok,
        "expected": row["unCertNo1"],
        "found": cert_info["full"]
    })

    if row["unCertNoShort"]:
        short_ok = cert_info["shortCert"] == str(row["unCertNoShort"]).strip()
        checks.append({
            "field": "Certificate ref.",
            "ok": short_ok,
            "expected": row["unCertNoShort"],
            "found": cert_info["shortCert"]
        })

    if cert_info["validUpto"]:
        try:
            valid_date = datetime.strptime(
                cert_info["validUpto"], "%d %B %Y"
            )
            expired = valid_date < datetime.now()
        except Exception:
            expired = False

        checks.append({
            "field": "Validity",
            "ok": not expired,
            "expected": "not expired",
            "found": cert_info["validUpto"] + (
                " (EXPIRED)" if expired else " (valid)"
            )
        })

    return checks


# ============================= GROUPING & AGGREGATION =============================

def packing_prefix(cert_str):
    s = str(cert_str or "").strip().upper()

    if s.startswith("4G"):
        return "BOX_4G"
    if s.startswith("1A1"):
        return "DRUM_1A1"
    if s.startswith("1A2"):
        return "DRUM_1A2"

    m = re.match(r"^[0-9][A-Z][0-9]?", s)
    return m.group(0) if m else s[:3]


def compute_group_key(row, threshold):
    fp = extract_number(row["flashPoint"])
    bucket = "LT" if fp is not None and fp < threshold else "GE"

    return "|".join([
        bucket,
        row["psn"].upper(),
        packing_prefix(row["unCertNo1"]),
        row["marinePollutant"],
        row["unNumber"]
    ])


def aggregate_group(rows):
    flash_values = [
        (extract_number(r["flashPoint"]), r["flashPoint"])
        for r in rows
        if extract_number(r["flashPoint"]) is not None
    ]
    min_flash = min(flash_values, key=lambda x: x[0])[1] if flash_values else ""

    pfx = packing_prefix(rows[0]["unCertNo1"])
    pkg_label = "DRUMS" if pfx.startswith("DRUM") else "BOXES"

    # Requested correction:
    # If outer package is DRUMS, INNER PACKING must not show.
    inner_packing = "" if pfx.startswith("DRUM") else (
        f"{sum_values([r['noInTins'] for r in rows])} TINS"
    )

    return {
        "rows": rows,
        "unNumber": rows[0]["unNumber"],
        "psn": rows[0]["psn"],
        "hazClass": rows[0]["hazClass"],
        "pkgGroup": rows[0]["pkgGroup"],
        "marinePollutant": rows[0]["marinePollutant"],
        "ems": ", ".join(uniq([r["ems"] for r in rows])),
        "technicalName": " , ".join(uniq([r["technicalName"] for r in rows])),
        "flashPoint": min_flash,
        "unPackingCode": " , ".join(uniq([r["unCertNo1"] for r in rows])),
        "outerPackages": f"{sum_values([r['noOfBoxes'] for r in rows])} {pkg_label}",
        "innerPacking": inner_packing,
        "grossWt": fmt_weight(sum_values([r["grossWt"] for r in rows])),
        "netWt": fmt_weight(sum_values([r["netWt"] for r in rows])),
        "materials": [r["materialDescription"] for r in rows],
    }


def compute_groups(rows, threshold):
    groups = {}

    for row in rows:
        key = compute_group_key(row, threshold)
        groups.setdefault(key, []).append(row)

    return [aggregate_group(rows) for rows in groups.values()]


# ============================= SHIPMENT DETAILS =============================

SHIPMENT_FIELDS = [
    ("shipperBlock", "Shipper", "Company name\nAddress line 1\nAddress line 2\nCountry"),
    ("consigneeBlock", "Consignee", "Company name\nAddress line 1\nAddress line 2\nCountry"),
    ("referenceNumbers", "Reference Number(s)", ""),
    ("emergencyContact", "24hr emergency contact", "Name – phone (with country code)"),
    ("carrier", "Carrier", ""),
    ("bookingNumber", "Carrier's booking number", "Fill in once booked"),
    ("shipNameVoyage", "Ship's name & voyage no.", "Fill in once booked"),
    ("containerNumbers", "Container number(s)", "Fill in once booked"),
    ("portOfLoading", "Port of loading", ""),
    ("portOfUnloading", "Port of unloading", ""),
    ("finalPlaceDelivery", "Final place of delivery", ""),
    ("placeDate", "Place and date", "e.g. VADODARA-DD-MM-YYYY"),
    ("signatoryCompany", "Signatory company/org", ""),
    ("packerSignatory", "Signature on behalf of packer", ""),
    ("shipperSignatory", "Signature on behalf of shipper", ""),
]


# ============================= TEMPLATE OUTPUT =============================

def set_cell(ws, cell_ref, value):
    ws[cell_ref] = "" if value is None else value


def populate_manual_template(template_bytes, group, shipment):
    """
    Populate MANUAL DGD1.xlsx while retaining its existing formatting/layout.

    Because the exact cell mapping is template-specific, this function first
    searches the template for the labels used in the supplied HTML output and
    writes the corresponding value into the adjacent cells. Existing template
    formatting/merged cells remain untouched.
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active

    # Search labels and write into the next useful cell.
    label_map = {
        "Reference Number(s)": shipment.get("referenceNumbers", ""),
        "Carrier": shipment.get("carrier", ""),
        "Carrier's Booking Number:": shipment.get("bookingNumber", ""),
        "Name/Status, Company/Organization of Signatory": shipment.get("signatoryCompany", ""),
        "Place and Date": shipment.get("placeDate", ""),
        "Signature on Behalf of Packer": shipment.get("packerSignatory", ""),
        "Signature on Behalf of Shipper": shipment.get("shipperSignatory", ""),
        "Port of Loading": shipment.get("portOfLoading", ""),
        "Port of Unloading": shipment.get("portOfUnloading", ""),
        "Final Place of Delivery": shipment.get("finalPlaceDelivery", ""),
        "Container Number(s)": shipment.get("containerNumbers", ""),
        "Ship's Name and Voyage No.": shipment.get("shipNameVoyage", ""),
        "UN NO": group["unNumber"],
        "PROPER SHIPPING NAME": group["psn"],
        "TECHNICAL  NAME": group["technicalName"],
        "TECHNICAL NAME": group["technicalName"],
        "CLASS": group["hazClass"],
        "UN PACKING GROUP": group["pkgGroup"],
        "UN PACKING CODE": group["unPackingCode"],
        "EMS NO.": group["ems"],
        "FLASH POINT": group["flashPoint"],
        "MARINE POLLUTANT": group["marinePollutant"],
        "GR WT": group["grossWt"],
        "NT WT": group["netWt"],
        "OUTER PACKAGES": group["outerPackages"],
        "INNER PACKING": group["innerPacking"],
    }

    # Split shipper/consignee blocks.
    shipper_lines = (shipment.get("shipperBlock") or "").splitlines()
    consignee_lines = (shipment.get("consigneeBlock") or "").splitlines()

    # Locate labels anywhere in the workbook.
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            label = str(cell.value).strip()
            upper = label.upper().rstrip(":- ").strip()

            # Exact/contains matching for the dynamic fields.
            matched_value = None
            for k, value in label_map.items():
                if k.upper().rstrip(":- ").strip() in upper or upper in k.upper():
                    matched_value = value
                    break

            if matched_value is not None:
                target = ws.cell(row=cell.row, column=min(cell.column + 1, ws.max_column))
                # Don't overwrite a label cell if it appears merged/occupied;
                # use next empty cell on the same row.
                if target.value not in (None, ""):
                    for c in range(cell.column + 1, ws.max_column + 1):
                        candidate = ws.cell(row=cell.row, column=c)
                        if candidate.value in (None, ""):
                            target = candidate
                            break
                target.value = matched_value

    # Handle shipper/consignee using the visible labels from the original.
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or "").strip().upper() == "SHIPPER":
                for offset, value in enumerate(shipper_lines, start=1):
                    ws.cell(row=cell.row + offset, column=cell.column).value = value
            if str(cell.value or "").strip().upper() == "CONSIGNEE":
                for offset, value in enumerate(consignee_lines, start=1):
                    ws.cell(row=cell.row + offset, column=cell.column).value = value

    # Requested correction: drums must have no inner packing.
    if str(group["outerPackages"]).upper().find("DRUM") >= 0:
        for row in ws.iter_rows():
            for cell in row:
                if "INNER PACKING" in str(cell.value or "").upper():
                    target = ws.cell(row=cell.row, column=min(cell.column + 1, ws.max_column))
                    target.value = ""

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================= SESSION STATE =============================

if "annexure_rows" not in st.session_state:
    st.session_state.annexure_rows = []
if "verification" not in st.session_state:
    st.session_state.verification = []
if "groups" not in st.session_state:
    st.session_state.groups = []
if "threshold" not in st.session_state:
    st.session_state.threshold = 23.0
if "shipment" not in st.session_state:
    st.session_state.shipment = {}
if "template_bytes" not in st.session_state:
    st.session_state.template_bytes = None


# ============================= STEP 1 =============================

st.markdown("### 1 · Upload source documents")

c1, c2, c3 = st.columns(3)

with c1:
    annexure_file = st.file_uploader(
        "Annexure — Required",
        type=["xlsx", "xls"],
        key="annexure"
    )

with c2:
    msds_files = st.file_uploader(
        "MSDS PDFs — Optional",
        type=["pdf"],
        accept_multiple_files=True,
        key="msds"
    )

with c3:
    cert_files = st.file_uploader(
        "UN Certificate PDFs — Optional",
        type=["pdf"],
        accept_multiple_files=True,
        key="certs"
    )

template_file = st.file_uploader(
    "MANUAL DGD1.xlsx — Final output template",
    type=["xlsx"],
    key="manual_template"
)

process = st.button(
    "Process files",
    type="primary",
    disabled=annexure_file is None
)

if process:
    try:
        annexure_bytes = annexure_file.getvalue()
        parsed = parse_annexure(annexure_bytes)

        if not parsed["rows"]:
            st.error(
                "No usable rows found — check the annexure has the required columns."
            )
        else:
            st.session_state.annexure_rows = parsed["rows"]

            msds_data = []
            for f in msds_files or []:
                b = f.getvalue()
                msds_data.append({
                    "name": f.name,
                    "text": extract_pdf_text(b)
                })

            cert_data = []
            for f in cert_files or []:
                b = f.getvalue()
                text = extract_pdf_text(b)
                cert_data.append({
                    "name": f.name,
                    "text": text,
                    "info": extract_cert_info(text)
                })

            verification = []

            for row in parsed["rows"]:
                msds_match = None
                msds_checks = []

                for msds in msds_data:
                    if match_msds_to_row(msds["text"], row):
                        msds_match = msds
                        break

                if msds_match:
                    msds_checks = verify_msds_against_row(
                        msds_match["text"], row
                    )

                cert_match = None
                cert_checks = []

                for cert in cert_data:
                    info = cert["info"]
                    if (
                        info
                        and row["unCertNoShort"]
                        and info["shortCert"] == str(row["unCertNoShort"]).strip()
                    ):
                        cert_match = cert
                        break

                if cert_match:
                    cert_checks = verify_cert_against_row(
                        cert_match["info"], row
                    )

                all_checks = msds_checks + cert_checks
                has_error = any(not x["ok"] for x in all_checks)
                has_warning = not msds_match or not cert_match

                status = "ERROR" if has_error else (
                    "WARNING" if has_warning else "OK"
                )

                verification.append({
                    "row": row,
                    "msdsMatch": msds_match,
                    "certMatch": cert_match,
                    "msdsChecks": msds_checks,
                    "certChecks": cert_checks,
                    "status": status
                })

            st.session_state.verification = verification
            st.session_state.groups = compute_groups(
                st.session_state.annexure_rows,
                st.session_state.threshold
            )

            if template_file:
                st.session_state.template_bytes = template_file.getvalue()

            st.success(
                f"Done — {len(parsed['rows'])} rows, "
                f"{len(msds_data)} MSDS, {len(cert_data)} certificate(s). "
                f"Parsed from \"{parsed['sheetName']}\" "
                f"({parsed['matchedCount']}/{parsed['totalRequired']} required columns matched)."
            )

    except Exception as e:
        st.exception(e)


# ============================= STEP 2 =============================

if st.session_state.verification:
    st.markdown("### 2 · Verification results")

    ok_count = sum(v["status"] == "OK" for v in st.session_state.verification)
    warn_count = sum(v["status"] == "WARNING" for v in st.session_state.verification)
    err_count = sum(v["status"] == "ERROR" for v in st.session_state.verification)

    st.caption(
        f"{ok_count} OK · {warn_count} warning · {err_count} error"
    )

    table_rows = []
    for v in st.session_state.verification:
        table_rows.append({
            "Material": v["row"]["materialDescription"],
            "UN No.": v["row"]["unNumber"],
            "MSDS": v["msdsMatch"]["name"] if v["msdsMatch"] else "— not uploaded",
            "UN Cert": v["certMatch"]["name"] if v["certMatch"] else "— not uploaded",
            "Status": v["status"]
        })

    st.dataframe(pd.DataFrame(table_rows), use_container_width=True)

    st.info(
        "You can proceed even if flags are raised. Declarations are generated "
        "from the annexure data — flags are for your review, not blockers."
    )


# ============================= STEP 3 =============================

if st.session_state.annexure_rows:
    st.markdown("### 3 · Grouping rules & shipment details")

    st.markdown("**Active grouping sequence**")
    st.write(
        " → ".join(
            f"CHECK {i+1}: {c['label']} ({c['cond']})"
            for i, c in enumerate(CHECKPOINTS)
        )
    )

    st.session_state.threshold = st.number_input(
        "Flash point threshold (°C)",
        value=float(st.session_state.threshold),
        step=0.1
    )

    if st.button("Re-run grouping"):
        st.session_state.groups = compute_groups(
            st.session_state.annexure_rows,
            st.session_state.threshold
        )

    st.markdown("#### Shipment & document details")

    shipment = {}
    for key, label, placeholder in SHIPMENT_FIELDS:
        shipment[key] = st.text_area(
            label,
            value=st.session_state.shipment.get(key, ""),
            placeholder=placeholder,
            key=f"shipment_{key}"
        )

    st.session_state.shipment = shipment

    if st.button("Generate declarations", type="primary"):
        st.session_state.groups = compute_groups(
            st.session_state.annexure_rows,
            st.session_state.threshold
        )


# ============================= STEP 4 =============================

if st.session_state.groups:
    st.markdown("### 4 · Generated declarations")
    st.caption(
        f"{len(st.session_state.groups)} declaration(s) from "
        f"{len(st.session_state.annexure_rows)} annexure rows"
    )

    generated_files = []

    for idx, g in enumerate(st.session_state.groups, start=1):
        with st.container(border=True):
            st.markdown(
                f"**DGD-{idx} · UN {g['unNumber']} · {g['psn']}**"
            )
            st.caption(
                f"{len(g['materials'])} product line(s): "
                + ", ".join(g["materials"])
            )

            a, b = st.columns(2)
            with a:
                st.write(f"Class / Packing group: {g['hazClass']} / {g['pkgGroup']}")
                st.write(f"Marine pollutant: {g['marinePollutant']}")
                st.write(f"Flash point: {g['flashPoint']}")
                st.write(f"EMS: {g['ems']}")
            with b:
                st.write(f"Outer packages: {g['outerPackages']}")
                st.write(f"Inner packing: {g['innerPacking'] or '—'}")
                st.write(f"Gross / Net weight: {g['grossWt']} / {g['netWt']}")
                st.write(f"UN packing code: {g['unPackingCode']}")

            if st.session_state.template_bytes:
                output_bytes = populate_manual_template(
                    st.session_state.template_bytes,
                    g,
                    st.session_state.shipment
                )
                generated_files.append(
                    (f"DGDI-{idx}.xlsx", output_bytes)
                )

                st.download_button(
                    f"Download DGDI-{idx}.xlsx",
                    data=output_bytes,
                    file_name=f"DGDI-{idx}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_{idx}"
                )
            else:
                st.warning(
                    "Upload MANUAL DGD1.xlsx to generate the final template-format file."
                )

    if generated_files:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as z:
            for filename, data in generated_files:
                z.writestr(filename, data)

        st.download_button(
            "Download all (.zip)",
            data=zip_buffer.getvalue(),
            file_name="DGDI-declarations.zip",
            mime="application/zip"
        )

st.caption(
    "This tool fills IMO DGD (IMDG) declaration files from your own annexure "
    "and source documents — it replaces manual data transfer, not your document "
    "formats. Review every generated declaration against your source documents "
    "before submission to the carrier."
)
